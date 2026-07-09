"""
pipeline.panel.avance_centros — Reporte de avance por centro (Excel consolidado).

Genera un Excel de una sola hoja con una fila por centro y las columnas:
  - Centro
  - Ingresos (n)
  - En tratamiento (n)
  - Egresos (n)
  - Seguimientos (n)
  - Total registros (n)
  - Pacientes con continuidad (n y %)
  - Último TOP (fecha)
  - Actividad (Verde / Amarillo / Rojo)

El archivo se descarga como botón desde el panel, sin necesidad de la base Wide.
Trabaja directamente sobre df_panel (datos en caché de Supabase).

Función expuesta:
  boton_descarga(df, pais, centro_id=None)
"""
import io
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
from openpyxl.utils import get_column_letter

from pipeline.panel.config import (
    COLOR_VERDE, COLOR_AMARILLO, COLOR_ROJO, COLOR_GRIS,
    actividad_por_centro, continuidad_por_centro, ingresos_por_centro,
    SEMAFORO_UMBRAL_VERDE, SEMAFORO_UMBRAL_AMARILLO,
)


# ─── Helpers de estilo openpyxl ──────────────────────────────────────────────

def _hex(color_str):
    """Convierte '#RRGGBB' → 'FFRRGGBB' (formato openpyxl)."""
    return 'FF' + color_str.lstrip('#').upper()


def _fill(hex_color):
    return PatternFill('solid', fgColor=_hex(hex_color))


def _font(bold=False, color='1F1F1F', size=10):
    return Font(bold=bold, color=_hex(color), size=size, name='Calibri')


def _border_thin():
    s = Side(style='thin', color='FFDDDDDD')
    return Border(left=s, right=s, top=s, bottom=s)


def _alinear(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ─── Construcción del DataFrame de avance ────────────────────────────────────

def _calcular_avance(df):
    """
    A partir del df completo del país calcula, por centro:
      ingresos, en_tratamiento, egresos, seguimientos, total,
      n_con_continuidad, pct_continuidad, ultima_fecha, dias, actividad.
    """
    if df is None or df.empty or 'centro' not in df.columns:
        return pd.DataFrame()

    tmp = df.copy()
    tmp['centro'] = tmp['centro'].astype(str).str.strip()
    tmp['etapa']  = tmp['etapa'].fillna('').astype(str).str.strip()
    tmp = tmp[tmp['centro'] != '']

    # Conteo por etapa
    etapas = ['ingreso', 'en_tratamiento', 'egreso', 'seguimiento']
    conteos = {}
    for etapa in etapas:
        conteos[etapa] = (
            tmp[tmp['etapa'] == etapa]
            .groupby('centro').size()
            .rename(etapa)
        )

    totales = tmp.groupby('centro').size().rename('total')

    # Continuidad
    cont = continuidad_por_centro(df).set_index('centro')

    # Actividad (días desde último registro)
    act = actividad_por_centro(df).set_index('centro')

    # Construir tabla
    centros = sorted(tmp['centro'].unique())
    filas = []
    for c in centros:
        ing   = int(conteos['ingreso'].get(c, 0))
        trat  = int(conteos['en_tratamiento'].get(c, 0))
        egr   = int(conteos['egreso'].get(c, 0))
        seg   = int(conteos['seguimiento'].get(c, 0))
        tot   = int(totales.get(c, 0))

        n_cont  = int(cont.loc[c, 'n_con_continuidad']) if c in cont.index else 0
        pct     = round(cont.loc[c, 'pct_continuidad'], 1) if c in cont.index else 0.0

        ult_f   = act.loc[c, 'ultima_fecha'] if c in act.index else pd.NaT
        dias    = act.loc[c, 'dias'] if c in act.index else None

        if dias is None or (isinstance(dias, float) and dias != dias):
            actividad = 'Sin datos'
        elif dias <= SEMAFORO_UMBRAL_VERDE:
            actividad = 'Verde'
        elif dias <= SEMAFORO_UMBRAL_AMARILLO:
            actividad = 'Amarillo'
        else:
            actividad = 'Rojo'

        filas.append({
            'Centro':             c,
            'Ingresos':           ing,
            'En tratamiento':     trat,
            'Egresos':            egr,
            'Seguimientos':       seg,
            'Total registros':    tot,
            'Con continuidad':    n_cont,
            '% Continuidad':      pct,
            'Último TOP':         ult_f.strftime('%d/%m/%Y') if pd.notna(ult_f) else '—',
            'Días sin registro':  int(dias) if dias is not None and dias == dias else '—',
            'Actividad':          actividad,
        })

    return pd.DataFrame(filas)


# ─── Construcción del Excel ───────────────────────────────────────────────────

_COLOR_ACTIVIDAD = {
    'Verde':    COLOR_VERDE,
    'Amarillo': COLOR_AMARILLO,
    'Rojo':     COLOR_ROJO,
    'Sin datos': COLOR_GRIS,
}

_COLUMNAS = [
    ('Centro',            18, 'left'),
    ('Ingresos',          10, 'center'),
    ('En tratamiento',    14, 'center'),
    ('Egresos',           10, 'center'),
    ('Seguimientos',      13, 'center'),
    ('Total registros',   14, 'center'),
    ('Con continuidad',   15, 'center'),
    ('% Continuidad',     14, 'center'),
    ('Último TOP',        13, 'center'),
    ('Días sin registro', 16, 'center'),
    ('Actividad',         12, 'center'),
]


def _generar_excel(df_avance, pais):
    """Genera el workbook openpyxl y lo devuelve como bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Avance por centro'

    # Encabezado del reporte
    ws.merge_cells('A1:K1')
    ws['A1'] = f'Reporte de avance por centro · {pais}'
    ws['A1'].font = _font(bold=True, color='1F3864', size=12)
    ws['A1'].alignment = _alinear('center')
    ws['A1'].fill = _fill('#D6E4F0')

    ws.merge_cells('A2:K2')
    ws['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font = _font(color='777777', size=9)
    ws['A2'].alignment = _alinear('center')

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 14

    # Encabezados de columnas (fila 3)
    HDR_FILL = _fill('#1F3864')
    HDR_FONT = _font(bold=True, color='FFFFFF', size=10)
    for col_idx, (header, ancho, alin) in enumerate(_COLUMNAS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = _alinear(alin, wrap=True)
        cell.border    = _border_thin()
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    ws.row_dimensions[3].height = 28

    # Filas de datos (desde fila 4)
    for row_idx, (_, row) in enumerate(df_avance.iterrows(), start=4):
        alt = row_idx % 2 == 0
        fondo = '#F7FAFC' if alt else '#FFFFFF'

        valores = [
            row['Centro'],
            row['Ingresos'],
            row['En tratamiento'],
            row['Egresos'],
            row['Seguimientos'],
            row['Total registros'],
            row['Con continuidad'],
            f"{row['% Continuidad']}%",
            row['Último TOP'],
            row['Días sin registro'],
            row['Actividad'],
        ]

        for col_idx, (val, (_, _, alin)) in enumerate(zip(valores, _COLUMNAS), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = _font(size=10)
            cell.alignment = _alinear(alin)
            cell.border    = _border_thin()

            # Colorear celda de Actividad
            if col_idx == len(_COLUMNAS):
                color_act = _COLOR_ACTIVIDAD.get(str(val), COLOR_GRIS)
                cell.fill = _fill(color_act)
                cell.font = _font(bold=True, color='FFFFFF', size=10)
            else:
                cell.fill = _fill(fondo)

        ws.row_dimensions[row_idx].height = 18

    # Fila de totales
    fila_tot = len(df_avance) + 4
    ws.merge_cells(f'A{fila_tot}:A{fila_tot}')
    ws.cell(fila_tot, 1, 'TOTAL').font      = _font(bold=True, size=10)
    ws.cell(fila_tot, 1, 'TOTAL').fill      = _fill('#D6E4F0')
    ws.cell(fila_tot, 1, 'TOTAL').alignment = _alinear('center')
    ws.cell(fila_tot, 1, 'TOTAL').border    = _border_thin()

    totales_cols = [2, 3, 4, 5, 6, 7]  # Ingresos → Con continuidad
    for col_idx in totales_cols:
        letra = get_column_letter(col_idx)
        cell  = ws.cell(fila_tot, col_idx,
                        value=f'=SUM({letra}4:{letra}{fila_tot - 1})')
        cell.font      = _font(bold=True, size=10)
        cell.fill      = _fill('#D6E4F0')
        cell.alignment = _alinear('center')
        cell.border    = _border_thin()

    # Rellenar columnas restantes de totales
    for col_idx in range(8, len(_COLUMNAS) + 1):
        cell = ws.cell(fila_tot, col_idx, value='')
        cell.fill   = _fill('#D6E4F0')
        cell.border = _border_thin()

    ws.freeze_panes = 'A4'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─── Función pública ──────────────────────────────────────────────────────────

def boton_descarga(df, pais, centro_id=None):
    """
    Renderiza el botón de descarga del Excel de avance por centro.
    Diseñado para insertarse dentro de un st.container(border=True) en app.py.

    Args:
        df:        DataFrame completo del país (df_panel).
        pais:      Nombre del país (str).
        centro_id: Si se pasa, filtra solo ese centro (modo drill-down v3.0).
    """
    if df is None or df.empty:
        st.caption('Sin datos disponibles para generar el reporte.')
        return

    df_avance = _calcular_avance(df)

    if df_avance.empty:
        st.caption('No se encontraron centros con datos.')
        return

    n_centros = len(df_avance)
    fecha_hoy = datetime.now().strftime('%Y%m%d')
    nombre_archivo = f'avance_{pais.lower().replace(" ", "_")}_{fecha_hoy}.xlsx'

    # Vista previa compacta
    st.markdown(
        f'<div style="font-size:.78rem;color:#555;margin-bottom:.4rem;">'
        f'  {n_centros} centros · datos al {datetime.now().strftime("%d/%m/%Y")}'
        f'</div>',
        unsafe_allow_html=True,
    )

    excel_bytes = _generar_excel(df_avance, pais)

    st.download_button(
        label='⬇️  Descargar Excel de avance',
        data=excel_bytes,
        file_name=nombre_archivo,
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        use_container_width=True,
    )
